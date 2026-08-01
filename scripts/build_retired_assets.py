from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.worksheet.properties import PageSetupProperties

INK = "1F2433"
MUTED = "8A8F9A"
ACCENT = "222D65"
HAIRLINE = "ECEDF0"
FONT = "Arial"
CEMENT = "Cement / Tanker Trailer"

HEADERS = ["Type", "Unit #", "Make", "Model", "Issues", "Est. Cost to Repair"]
# Type, Unit, Make, Model, Issues
ASSETS = [
    ("Mixer", "149", "Kenworth", "", ""),
    ("Mixer", "29", "Kenworth", "", "Wrecked, parting out — multiple issues; will not repair."),
    ("Tractor", "D94", "Peterbilt", "", ""),
    ("End Dump Trailer", "—", "Travis", "", "No unit number assigned."),
    ("End Dump Trailer", "D005", "", "", ""),
    ("Tractor", "C106", "", "", "Wrecked — missing hood."),
    (CEMENT, "FA-106", "", "", ""),
    ("Tractor", "C502", "Sterling", "", ""),
    (CEMENT, "FA002", "", "", ""),
    (CEMENT, "C-101", "", "", ""),
    (CEMENT, "C007", "", "", ""),
    (CEMENT, "C001", "", "", ""),
]

wb = Workbook()
ws = wb.active
ws.title = "Retired Assets"
ws.sheet_view.showGridLines = False
for col, w in {"A": 25, "B": 10, "C": 14, "D": 12, "E": 42, "F": 17}.items():
    ws.column_dimensions[col].width = w

ws.row_dimensions[1].height = 24
t = ws.cell(1, 1, "Retired Assets — Repair Assessment")
t.font = Font(name=FONT, size=14, bold=True, color=INK)
t.alignment = Alignment(horizontal="left", vertical="center")
ws.row_dimensions[2].height = 8

left = Alignment(horizontal="left", vertical="center", indent=1)
center = Alignment(horizontal="center", vertical="center")
right = Alignment(horizontal="right", vertical="center", indent=1)
navy_rule = Side(style="medium", color=ACCENT)
hair = Side(style="thin", color=HAIRLINE)

HEAD = 3
ws.row_dimensions[HEAD].height = 22
LEFT_COLS = {1, 3, 4, 5}  # Type, Make, Model, Issues
for idx, label in enumerate(HEADERS, start=1):
    c = ws.cell(HEAD, idx, label.upper())
    c.font = Font(name=FONT, size=9, bold=True, color=MUTED)
    c.alignment = left if idx in LEFT_COLS else (right if idx == 6 else center)
    c.border = Border(bottom=navy_rule)

FIRST = HEAD + 1
for i, (atype, unit, make, model, issues) in enumerate(ASSETS):
    r = FIRST + i
    ws.row_dimensions[r].height = 24
    a = ws.cell(r, 1, atype); a.font = Font(name=FONT, size=11, color=INK); a.alignment = left
    b = ws.cell(r, 2, unit)
    b.font = Font(name=FONT, size=11, color=MUTED if unit == "—" else INK); b.alignment = center
    mk = ws.cell(r, 3, make); mk.font = Font(name=FONT, size=11, color=INK); mk.alignment = left
    md = ws.cell(r, 4, model); md.font = Font(name=FONT, size=11, color=INK); md.alignment = left
    iss = ws.cell(r, 5, issues)
    iss.font = Font(name=FONT, size=10, color=INK if issues else MUTED, italic=not issues)
    iss.alignment = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True)
    co = ws.cell(r, 6); co.number_format = '$#,##0'; co.font = Font(name=FONT, size=11, color=INK)
    co.alignment = right
    for col in range(1, 7):
        ws.cell(r, col).border = Border(bottom=hair)

ws.freeze_panes = "A4"
ws.print_title_rows = "1:3"
ws.page_setup.orientation = "landscape"
ws.page_setup.fitToWidth = 1
ws.page_setup.fitToHeight = 0
ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
ws.page_margins.left = ws.page_margins.right = 0.4
ws.page_margins.top = ws.page_margins.bottom = 0.5

wb.save("Retired_Assets_Repair_Assessment.xlsx")
print("saved")
